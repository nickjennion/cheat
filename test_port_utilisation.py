import openpyxl
import pytest

from port_utilisation import (
    write_utilisation_sheet,
    write_summary_excel,
    analyse_workbook,
)


# ---------------------------------------------------------------------------
# write_utilisation_sheet — hardware breakdown columns (G onwards)
# ---------------------------------------------------------------------------

def test_utilisation_sheet_has_six_stack_member_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    results = {"sw-a": (3, 1)}
    write_utilisation_sheet(ws, results, 42, hardware={"sw-a": {1: "C9300-48P"}})
    # Columns G..L (7..12)
    headers = [ws.cell(row=1, column=c).value for c in range(7, 13)]
    assert headers == [f"Stack Member {i}" for i in range(1, 7)]


def test_utilisation_sheet_populates_hardware_per_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    results = {"sw-a": (3, 1)}
    hardware = {"sw-a": {1: "C9300-48P", 2: "C9300-24P"}}
    write_utilisation_sheet(ws, results, 42, hardware=hardware)
    # data row is row 2 (row 1 headers); sorted single switch
    assert ws.cell(row=2, column=7).value == "C9300-48P"   # Stack Member 1
    assert ws.cell(row=2, column=8).value == "C9300-24P"   # Stack Member 2
    assert ws.cell(row=2, column=9).value in (None, "")     # Stack Member 3 blank


def test_utilisation_sheet_hardware_optional():
    wb = openpyxl.Workbook()
    ws = wb.active
    results = {"sw-a": (3, 1)}
    # No hardware supplied — headers still present, cells blank
    write_utilisation_sheet(ws, results, 42)
    assert ws.cell(row=1, column=7).value == "Stack Member 1"
    assert ws.cell(row=2, column=7).value in (None, "")


# ---------------------------------------------------------------------------
# write_summary_excel — same breakdown in the standalone export
# ---------------------------------------------------------------------------

def test_summary_excel_writes_hardware(tmp_path):
    out = tmp_path / "summary.xlsx"
    results = {"sw-a": (2, 2)}
    hardware = {"sw-a": {1: "C9300-48P"}}
    ok, _ = write_summary_excel(results, 42, str(out), hardware=hardware)
    assert ok
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=7).value == "Stack Member 1"
    assert ws.cell(row=2, column=7).value == "C9300-48P"


# ---------------------------------------------------------------------------
# analyse_workbook — collect member -> model from Stack Member / Model columns
# ---------------------------------------------------------------------------

def test_analyse_workbook_collects_hardware(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sw-a"
    for col, h in enumerate(
        ["Switch", "Stack Member", "Model", "Interface", "Last Input"], start=1
    ):
        ws.cell(row=1, column=col, value=h)
    rows = [
        ("sw-a", "1", "C9300-48P", "Gi1/0/1", "2d"),
        ("sw-a", "2", "C9300-24P", "Gi2/0/1", "never"),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    path = tmp_path / "report.xlsx"
    wb.save(path)

    success, _msg, results, hardware = analyse_workbook(str(path), 42)
    assert success
    assert hardware["sw-a"] == {1: "C9300-48P", 2: "C9300-24P"}
