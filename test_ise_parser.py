"""Unit tests for ise_parser and the ISE endpoint Excel/CSV writers."""

from types import SimpleNamespace


def _res(**kw):
    base = dict(id="E1", name="ep-1", mac="0011.2233.4455", description="camera",
                profileId="P1", groupId="G1", portalUser="jane",
                staticGroupAssignment=True, staticProfileAssignment=False)
    base.update(kw)
    return SimpleNamespace(**base)


def test_parse_endpoint_normalises_fields():
    from ise_parser import parse_endpoint
    e = parse_endpoint(_res())
    assert e.name == "ep-1"
    assert e.mac == "0011.2233.4455"
    assert e.profile_id == "P1"
    assert e.group_id == "G1"
    assert e.portal_user == "jane"
    assert e.static_group == "yes"
    assert e.static_profile == "no"


def test_parse_endpoints_resolves_group_names():
    from ise_parser import parse_endpoints
    eps = parse_endpoints([_res(groupId="G1"), _res(id="E2", groupId="G9")],
                          group_names={"G1": "CCTV Cameras"})
    assert eps[0].group_name == "CCTV Cameras"
    assert eps[1].group_name == "G9"  # unknown id falls back to the raw id


def test_parse_endpoints_tolerates_missing_attributes():
    from ise_parser import parse_endpoints
    eps = parse_endpoints([SimpleNamespace(id="E1")])
    assert eps[0].name == ""
    assert eps[0].static_group == ""
    assert eps[0].group_name == ""


def test_write_ise_endpoint_excel_and_csv(tmp_path):
    import csv
    import openpyxl
    from ise_parser import parse_endpoints
    from excel_generator import (
        write_ise_endpoint_excel,
        write_ise_endpoint_csv,
        ISE_ENDPOINT_HEADERS,
    )

    eps = parse_endpoints([_res()], group_names={"G1": "CCTV Cameras"})

    xlsx = tmp_path / "ise.xlsx"
    ok, msg = write_ise_endpoint_excel(eps, str(xlsx))
    assert ok
    ws = openpyxl.load_workbook(xlsx)["ISE Endpoints"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(ISE_ENDPOINT_HEADERS) + 1)]
    assert headers == ISE_ENDPOINT_HEADERS
    assert ws.cell(row=2, column=5).value == "CCTV Cameras"

    cpath = tmp_path / "ise.csv"
    ok, _ = write_ise_endpoint_csv(eps, str(cpath))
    assert ok
    with open(cpath, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ISE_ENDPOINT_HEADERS
    assert len(rows) == 2


def test_write_ise_endpoint_excel_refuses_empty(tmp_path):
    from excel_generator import write_ise_endpoint_excel
    ok, msg = write_ise_endpoint_excel([], str(tmp_path / "ise.xlsx"))
    assert ok is False
    assert "No ISE endpoints" in msg
