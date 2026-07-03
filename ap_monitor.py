"""
Access Point — Monitor Physical Movements.

Filter/select Unified APs, then display a live-refreshed table comparing
previous upstream (Assurance events, 24h) vs current upstream (physical topology).
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Sentinel display strings
_NO_CHANGE = "N/A — No Change"
_OFFLINE   = "— (offline)"
_NO_DATA   = "— (no data)"
_ERROR     = "— (error)"

_SENTINELS = (_OFFLINE, _ERROR, _NO_DATA, _NO_CHANGE)

# Column widths for the results table
_W_HOST = 42
_W_UP   = 20
_W_CDP  = 38


# ============================================================================
# Data Layer
# ============================================================================

def build_table_rows(
    selected_aps: list[dict],
    topology: dict[str, Optional[str]],
    events: dict[str, Optional[str]],
    topology_error: bool = False,
    events_error: bool = False,
) -> list[dict]:
    """Build display rows from AP dicts and upstream lookup dicts.

    Each row: {"hostname": str, "uptime": str, "previous": str, "current": str}
    """
    rows = []
    for ap in selected_aps:
        ap_id = ap.get("id", "")
        hostname = str(ap.get("hostname") or "unknown")
        uptime = str(ap.get("upTime") or "—")

        if topology_error:
            current = _ERROR
        else:
            val = topology.get(ap_id)
            current = val if val is not None else _OFFLINE

        if events_error:
            previous = _ERROR
        else:
            val = events.get(ap_id)
            previous = val if val is not None else _NO_DATA

        # No change: both valid, same value
        if (current not in _SENTINELS
                and previous not in _SENTINELS
                and current == previous):
            current = _NO_CHANGE
            previous = _NO_CHANGE

        rows.append({
            "hostname": hostname,
            "uptime": uptime,
            "previous": previous,
            "current": current,
        })
    return rows


def _ap_matches(ap: dict, filter_terms: list[str], exclude_terms: list[str]) -> bool:
    """Return True if AP matches all include terms and no exclude terms.

    '|' within a term = OR alternatives. Matches hostname and platformId.
    """
    text = (
        (ap.get("hostname") or "") + " " + (ap.get("platformId") or "")
    ).lower()
    for term in filter_terms:
        alts = [a.strip() for a in term.split("|") if a.strip()]
        if not any(a in text for a in alts):
            return False
    for term in exclude_terms:
        alts = [a.strip() for a in term.split("|") if a.strip()]
        if any(a in text for a in alts):
            return False
    return True


def _filter_label(filter_terms: list[str], exclude_terms: list[str]) -> str:
    """Return a human-readable summary of the active filter."""
    parts = []
    if filter_terms:
        parts.append("  AND  ".join(f"[{t}]" for t in filter_terms))
    if exclude_terms:
        parts.append("NOT " + "  NOT ".join(f"[{t}]" for t in exclude_terms))
    return "  ".join(parts) if parts else "(none)"


def _parse_numbers(entry: str, max_idx: int) -> list[int]:
    """Parse a comma-separated list of numbers and ranges into a sorted list.

    Numbers outside [1, max_idx] are silently dropped.
    Example: "1,3-5,7" with max_idx=6 → [1, 3, 4, 5]
    """
    result = set()
    for part in entry.split(","):
        part = part.strip()
        if "-" in part:
            try:
                lo, hi = part.split("-", 1)
                result.update(range(int(lo), int(hi) + 1))
            except ValueError:
                pass
        elif part.isdigit():
            result.add(int(part))
    return sorted(i for i in result if 1 <= i <= max_idx)


# ============================================================================
# UI Helpers
# ============================================================================

def _clear() -> None:
    import os
    os.system("clear" if os.name != "nt" else "cls")


def _pause() -> None:
    input("\n  Press Enter to continue...")


# ============================================================================
# Filter / Select Screen
# ============================================================================

def filter_select_screen(aps: list[dict]) -> list[dict]:
    """Filter and select APs. Returns selected dicts, or [] if user pressed b."""
    filter_terms: list[str] = []
    exclude_terms: list[str] = []
    selected: set[int] = set()

    while True:
        filtered = [
            ap for ap in aps
            if _ap_matches(ap, filter_terms, exclude_terms)
        ] if (filter_terms or exclude_terms) else []

        _clear()
        print("  Access Point — Monitor Physical Movements\n")
        print(f"  Filters: {_filter_label(filter_terms, exclude_terms)}\n")

        if not filter_terms and not exclude_terms:
            print("  Add at least one filter to show matching APs.")
            print("  Use '|' for OR within a term  (e.g. f bldg-a|bldg-b)")
        elif not filtered:
            print("  No APs matched — try 'fc' to clear filters and start over.")
        else:
            print(f"  {'#':<5} {'':3} {'AP Hostname':<42} {'Model':<22} {'IP Address'}")
            print(f"  {'-'*5} {'-'*3} {'-'*42} {'-'*22} {'-'*15}")
            for i, ap in enumerate(filtered, 1):
                check = "[X]" if i in selected else "[ ]"
                h = str(ap.get("hostname") or "unknown")
                p = str(ap.get("platformId") or "")
                ip = str(ap.get("managementIpAddress") or "")
                print(f"  {i:<5} {check} {h:<42} {p:<22} {ip}")
            print(f"\n  Selected: {len(selected)} AP(s)")

        print()
        print("  'f <term>'  add filter (| = OR)  |  'r <term>'  exclude  |  'fc'  clear all")
        print("  number(s)   toggle selection (e.g. 1  or  1,3-5)")
        print("  'p'  Proceed    |    'b'  Back")
        print()
        entry = input("  > ").strip()

        if entry.lower() == "b":
            return []
        elif entry.lower() in ("p", ""):
            if not selected:
                print("\n  Select at least one AP first.")
                _pause()
            else:
                return [filtered[i - 1] for i in sorted(selected) if i <= len(filtered)]
        elif entry.lower() == "fc":
            filter_terms.clear()
            exclude_terms.clear()
            selected.clear()
        elif entry.lower().startswith("f "):
            term = entry[2:].strip().lower()
            if term:
                filter_terms.append(term)
                selected.clear()
        elif entry.lower().startswith("r "):
            term = entry[2:].strip().lower()
            if term:
                exclude_terms.append(term)
                selected.clear()
        else:
            indices = _parse_numbers(entry, len(filtered))
            if not indices:
                print("\n  Unrecognised input.")
                _pause()
            else:
                for idx in indices:
                    if idx in selected:
                        selected.discard(idx)
                    else:
                        selected.add(idx)


# ============================================================================
# Results Table
# ============================================================================

def _print_table(rows: list[dict]) -> None:
    header = (
        f"  {'AP Hostname':<{_W_HOST}} {'Uptime':<{_W_UP}} "
        f"{'Previous Upstream (24h)':<{_W_CDP}} {'Current Upstream':<{_W_CDP}}"
    )
    sep = f"  {'-'*_W_HOST} {'-'*_W_UP} {'-'*_W_CDP} {'-'*_W_CDP}"
    print(header)
    print(sep)
    for row in rows:
        moved = (
            row["previous"] not in (_NO_CHANGE, _NO_DATA, _ERROR, _OFFLINE)
            and row["current"] not in (_NO_CHANGE, _ERROR, _OFFLINE)
            and row["previous"] != row["current"]
        )
        marker = " *" if moved else "  "
        print(
            f"{marker} {row['hostname']:<{_W_HOST}} {row['uptime']:<{_W_UP}} "
            f"{row['previous']:<{_W_CDP}} {row['current']:<{_W_CDP}}"
        )


def _export_excel(rows: list[dict], stem: str) -> None:
    """Write AP movement table to Excel in excel_reports/."""
    excel_dir = Path("excel_reports").resolve()
    excel_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d-%H-%M")
    outpath = str(excel_dir / f"{stem}-{ts}.xlsx")

    headers = ["AP Hostname", "Uptime", "Previous Upstream (24h)", "Current Upstream"]
    col_widths = [_W_HOST, _W_UP, _W_CDP, _W_CDP]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AP Movements"

    hdr_font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    hdr_fill = PatternFill("solid", start_color="FF2B579A")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="FFB0B0B0"),
        right=Side(style="thin", color="FFB0B0B0"),
    )
    dat_font = Font(name="Arial", size=10)
    dat_align = Alignment(vertical="center")
    moved_fill = PatternFill("solid", start_color="FFFFF3CD")   # amber — AP moved
    same_fill  = PatternFill("solid", start_color="FFD4EDDA")   # green  — no change
    other_fill = PatternFill("solid", start_color="FFFFFFFF")   # white  — error/offline/no-data

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        if row["current"] == _NO_CHANGE:
            fill = same_fill
        elif (
            row["current"] not in (_ERROR, _OFFLINE)
            and row["previous"] not in (_NO_DATA, _ERROR, _NO_CHANGE)
        ):
            fill = moved_fill
        else:
            fill = other_fill

        for col, val in enumerate(
            [row["hostname"], row["uptime"], row["previous"], row["current"]], start=1
        ):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = dat_font
            cell.alignment = dat_align
            cell.border = thin
            cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    wb.save(outpath)
    print(f"\n  ✓ Saved: {outpath}")


# ============================================================================
# Results Screen
# ============================================================================

def results_screen(client, selected_aps: list[dict]) -> None:
    """Show AP movement table. Keys: r=refresh, e=export, b=back."""
    ap_ids = [ap["id"] for ap in selected_aps]
    stem = "ap-monitor"

    while True:
        print("\n  Fetching current topology...", flush=True)
        topology, topology_error = client.get_ap_topology(ap_ids)

        print("  Fetching Assurance events (24h)...", flush=True)
        events, events_error = client.get_ap_events(ap_ids, hours=24)

        rows = build_table_rows(selected_aps, topology, events, topology_error, events_error)

        _clear()
        print("  Access Point — Monitor Physical Movements\n")
        print(f"  APs monitored: {len(selected_aps)}   * = upstream changed\n")
        _print_table(rows)
        print()
        if topology_error:
            print("  ⚠ Topology fetch failed — current upstream unavailable.")
        if events_error:
            print("  ⚠ Events fetch failed — previous upstream unavailable.")
        print()
        print("  'r' Refresh   'e' Export to Excel   'b' Back")
        print()
        key = input("  > ").strip().lower()

        if key == "b":
            return
        elif key == "r":
            _clear()
            print("  Refreshing...\n")
        elif key == "e":
            _export_excel(rows, stem)
            _pause()


# ============================================================================
# Entry Point
# ============================================================================

def run(client, aps: list[dict]) -> None:
    """Entry point: filter/select screen → results screen loop."""
    if not aps:
        print("\n  No Unified APs found in DNAC inventory.")
        _pause()
        return

    while True:
        selected = filter_select_screen(aps)
        if not selected:
            return
        results_screen(client, selected)
