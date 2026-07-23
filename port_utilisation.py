"""
Analyse port utilisation from a CHEAT UNPLUGGED report.

Reads an Excel workbook (per-stack tabs or consolidated single sheet) and
counts ports with traffic in the last 42 days vs. idle ports per switch/stack.

The "Last Input" column drives the classification:
  - "never" or missing → idle (no traffic ever)
  - Recent time (e.g., "2d3h", "5w") → in use (within threshold)
  - Old time (e.g., "10w") → idle (beyond threshold)

Usage:
    python port_utilisation.py <report.xlsx> [threshold_days]

Default threshold is 42 days (matching uptime highlight logic).
Output includes both stdout summary and timestamped Excel file.
"""

import sys
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from time_utils import parse_duration_days


# Hardware breakdown columns appended to the utilisation table (column G onwards).
HARDWARE_START_COL = 7
HARDWARE_MEMBER_COUNT = 6
HARDWARE_END_COL = HARDWARE_START_COL + HARDWARE_MEMBER_COUNT - 1  # column L
HARDWARE_HEADERS = [f"Stack Member {i}" for i in range(1, HARDWARE_MEMBER_COUNT + 1)]


def _member_to_int(value) -> Optional[int]:
    """Coerce a stack-member cell value ('1', 1, ' 2 ') to int, else None."""
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _write_hardware_headers(ws, header_font, header_fill, header_align) -> None:
    """Write the six 'Stack Member N' headers starting at column G."""
    for i, header in enumerate(HARDWARE_HEADERS):
        col = HARDWARE_START_COL + i
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_hardware_row(ws, row: int, member_models: dict) -> None:
    """Write per-member models across columns G..L for one data row.

    member_models maps member number (int) -> model string.
    """
    for i in range(HARDWARE_MEMBER_COUNT):
        col = HARDWARE_START_COL + i
        ws.cell(row=row, column=col, value=member_models.get(i + 1, ""))


def parse_last_input_days(last_input_str: str) -> Optional[float]:
    """Convert a "Last Input" string to days ago, or None if invalid/never.

    Handles multiple formats:
      "2d3h" → 2.125 days
      "5w" → 35 days
      "222h" → 9.25 days
      "00:00:13" (HH:MM:SS) → ~0.00015 days (13 seconds)
      "00:00:11" → ~0.00013 days (11 seconds)
      "never" or "" → None
    """
    return parse_duration_days(str(last_input_str).strip())


def is_copper_port(iface: str) -> bool:
    """Check if interface is a base copper port.

    Matches stacked format (Gi1/0/1, Te1/0/1) and non-stacked format
    (Gi0/1, Fa0/1) used by older switches such as the 3560.
    """
    if not iface:
        return False
    iface_upper = str(iface).upper()
    return bool(
        re.match(r"^(GI|TE)\d+/0/\d+$", iface_upper) or  # stacked: Gi1/0/1
        re.match(r"^(GI|FA)\d+/\d+$", iface_upper)         # non-stacked: Gi0/1, Fa0/1
    )


def analyse_workbook(
    wb_path: str, threshold_days: int = 42
) -> tuple[bool, str, dict]:
    """Analyse port utilisation from Excel workbook.

    Only counts base copper ports (GiX/0/X and TeX/0/X).
    Returns (success, message, results: dict[switch, (in_use, idle)],
             hardware: dict[switch, dict[member_num, model]])
    """
    in_path = Path(wb_path).resolve()
    if not in_path.is_file():
        return False, f"✗ File not found: {wb_path}", {}, {}

    print(f"Loading '{wb_path}'...")
    try:
        wb = openpyxl.load_workbook(in_path, data_only=True)
    except Exception as e:
        return False, f"✗ Failed to open: {e}", {}, {}

    results: dict[str, tuple[int, int]] = {}  # switch → (in_use, idle)
    hardware: dict[str, dict[int, str]] = {}  # switch → {member_num: model}

    # Known summary sheets produced by write_combined_excel — skip to avoid
    # double-counting ("All Ports" has the same Switch/Interface/Last Input
    # columns as per-stack sheets and would cause every port to be tallied twice).
    SKIP_TITLES = {"All Ports", "Port Utilisation"}

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        print(f"  Sheet {sheet_idx}/{len(wb.worksheets)}: '{ws.title}'...", end=" ", flush=True)

        if ws.title in SKIP_TITLES:
            print("(summary sheet, skipped)")
            continue

        # Find column indices by header
        if ws.max_row < 1:
            print("(empty)")
            continue

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            switch_col = headers.index("Switch") + 1
            iface_col = headers.index("Interface") + 1
            last_input_col = headers.index("Last Input") + 1
        except ValueError as e:
            print(f"✗ missing column ({e})")
            continue

        # Stack Member / Model columns drive the hardware breakdown (optional).
        member_col = headers.index("Stack Member") + 1 if "Stack Member" in headers else None
        model_col = headers.index("Model") + 1 if "Model" in headers else None

        # Tally ports per switch
        switch_stats: dict[str, tuple[int, int]] = {}
        row_count = 0
        skipped = 0

        for row in range(2, ws.max_row + 1):
            switch = ws.cell(row=row, column=switch_col).value
            iface = ws.cell(row=row, column=iface_col).value
            last_input = ws.cell(row=row, column=last_input_col).value

            if not switch:
                continue

            switch_str = str(switch).strip()

            # Record member -> model for the hardware breakdown (any port type).
            if member_col and model_col:
                member = _member_to_int(ws.cell(row=row, column=member_col).value)
                model = ws.cell(row=row, column=model_col).value
                if member is not None and model:
                    hardware.setdefault(switch_str, {}).setdefault(member, str(model).strip())

            # Only count base copper ports (Gi*/0/* and Te*/0/*)
            if not is_copper_port(iface):
                skipped += 1
                continue

            row_count += 1

            # Classify port as in-use or idle
            days_since_traffic = parse_last_input_days(last_input)
            in_use = days_since_traffic is not None and days_since_traffic < threshold_days
            is_in_use = 1 if in_use else 0
            is_idle = 0 if in_use else 1

            # Accumulate per switch
            if switch_str not in switch_stats:
                switch_stats[switch_str] = (0, 0)
            curr_in_use, curr_idle = switch_stats[switch_str]
            switch_stats[switch_str] = (curr_in_use + is_in_use, curr_idle + is_idle)

        # Merge into global results
        for switch, (in_use, idle) in switch_stats.items():
            if switch in results:
                prev_in_use, prev_idle = results[switch]
                results[switch] = (prev_in_use + in_use, prev_idle + idle)
            else:
                results[switch] = (in_use, idle)

        if skipped > 0:
            print(f"({row_count} copper ports across {len(switch_stats)} switch(es), {skipped} non-copper skipped)")
        else:
            print(f"({row_count} copper ports across {len(switch_stats)} switch(es))")

    wb.close()

    if not results:
        return False, "✗ No copper port data found", {}, {}

    return True, f"✓ Analysed {sum(in_use + idle for in_use, idle in results.values())} copper ports", results, hardware


def print_summary(results: dict[str, tuple[int, int]], threshold_days: int) -> None:
    """Print a nicely formatted summary table."""
    print(f"\n{'='*80}")
    print(f"Port Utilisation Summary (threshold: {threshold_days} days)")
    print(f"{'='*80}\n")

    print(f"{'Switch/Stack':<40} {'In Use':<12} {'Idle':<12} {'Total':<12}")
    print("-" * 80)

    grand_in_use = 0
    grand_idle = 0

    for switch in sorted(results.keys()):
        in_use, idle = results[switch]
        total = in_use + idle
        grand_in_use += in_use
        grand_idle += idle

        pct_in_use = (in_use / total * 100) if total > 0 else 0
        print(
            f"{switch:<40} {in_use:<12} {idle:<12} {total:<12} "
            f"({pct_in_use:.1f}% in use)"
        )

    print("-" * 80)
    grand_total = grand_in_use + grand_idle
    grand_pct = (grand_in_use / grand_total * 100) if grand_total > 0 else 0
    print(
        f"{'TOTAL':<40} {grand_in_use:<12} {grand_idle:<12} {grand_total:<12} "
        f"({grand_pct:.1f}% in use)"
    )
    print(f"{'='*80}\n")


def write_summary_excel(
    results: dict[str, tuple[int, int]], threshold_days: int, output_path: str | None = None,
    hardware: dict | None = None
) -> tuple[bool, str]:
    """Write port utilisation summary to an Excel file.

    hardware (optional) maps switch -> {member_num: model}; when supplied it
    populates the six 'Stack Member N' columns (G..L) at the row level.

    Returns (success: bool, message: str with file path if successful)
    """
    hardware = hardware or {}
    if output_path is None:
        stamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
        excel_dir = Path("excel_reports").resolve()
        if excel_dir.is_dir():
            output_path = str(excel_dir / f"port_utilisation_summary_{stamp}.xlsx")
        else:
            output_path = f"port_utilisation_summary_{stamp}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Headers
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
        header_fill = openpyxl.styles.PatternFill("solid", start_color="FF2B579A")
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        headers = ["Switch/Stack", "In Use", "Idle", "Total", "% In Use", "Threshold (days)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16

        _write_hardware_headers(ws, header_font, header_fill, header_align)

        # Data rows
        grand_in_use = 0
        grand_idle = 0
        row = 2

        for switch in sorted(results.keys()):
            in_use, idle = results[switch]
            total = in_use + idle
            grand_in_use += in_use
            grand_idle += idle
            pct = (in_use / total * 100) if total > 0 else 0

            ws.cell(row=row, column=1, value=switch)
            ws.cell(row=row, column=2, value=in_use)
            ws.cell(row=row, column=3, value=idle)
            ws.cell(row=row, column=4, value=total)
            ws.cell(row=row, column=5, value=pct / 100)
            ws.cell(row=row, column=6, value=threshold_days)

            # Format percentage column
            ws.cell(row=row, column=5).number_format = "0.0%"

            _write_hardware_row(ws, row, hardware.get(switch, {}))

            row += 1

        # Total row
        grand_total = grand_in_use + grand_idle
        grand_pct = (grand_in_use / grand_total * 100) if grand_total > 0 else 0

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=grand_in_use)
        ws.cell(row=row, column=3, value=grand_idle)
        ws.cell(row=row, column=4, value=grand_total)
        ws.cell(row=row, column=5, value=grand_pct / 100)
        ws.cell(row=row, column=6, value=threshold_days)

        # Bold and shade the total row
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = openpyxl.styles.Font(bold=True, name="Arial", size=10)
            cell.fill = openpyxl.styles.PatternFill("solid", start_color="FFE2E2E2")

        ws.cell(row=row, column=5).number_format = "0.0%"

        # Add borders to all cells
        thin_border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(style="thin", color="FFB0B0B0"),
            right=openpyxl.styles.Side(style="thin", color="FFB0B0B0"),
        )
        for r in range(1, row + 1):
            for c in range(1, HARDWARE_END_COL + 1):
                ws.cell(row=r, column=c).border = thin_border

        ws.freeze_panes = "A2"
        wb.save(output_path)
        return True, f"✓ Summary written to {output_path}"

    except Exception as e:
        return False, f"✗ Failed to write Excel: {e}"


def write_utilisation_sheet(ws, results: dict, threshold_days: int, hardware: dict | None = None) -> int:
    """Write port utilisation summary to an existing openpyxl worksheet.

    hardware (optional) maps switch -> {member_num: model}; when supplied it
    populates the six 'Stack Member N' columns (G..L) at the row level.
    """
    hardware = hardware or {}
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    header_fill = openpyxl.styles.PatternFill("solid", start_color="FF2B579A")
    header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    headers = ["Switch/Stack", "In Use", "Idle", "Total", "% In Use", "Threshold (days)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    col_widths = {"A": 40, "B": 12, "C": 12, "D": 12, "E": 14, "F": 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    _write_hardware_headers(ws, header_font, header_fill, header_align)

    grand_in_use = 0
    grand_idle = 0
    row = 2

    for switch in sorted(results.keys()):
        in_use, idle = results[switch]
        total = in_use + idle
        grand_in_use += in_use
        grand_idle += idle
        pct = (in_use / total) if total > 0 else 0.0

        ws.cell(row=row, column=1, value=switch)
        ws.cell(row=row, column=2, value=in_use)
        ws.cell(row=row, column=3, value=idle)
        ws.cell(row=row, column=4, value=total)
        pct_cell = ws.cell(row=row, column=5, value=pct)
        pct_cell.number_format = "0.0%"
        ws.cell(row=row, column=6, value=threshold_days)
        _write_hardware_row(ws, row, hardware.get(switch, {}))
        row += 1

    grand_total = grand_in_use + grand_idle
    grand_pct = (grand_in_use / grand_total) if grand_total > 0 else 0.0

    for col, val in enumerate(
        ["TOTAL", grand_in_use, grand_idle, grand_total, grand_pct, threshold_days], start=1
    ):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = openpyxl.styles.Font(bold=True, name="Arial", size=10)
        cell.fill = openpyxl.styles.PatternFill("solid", start_color="FFE2E2E2")

    ws.cell(row=row, column=5).number_format = "0.0%"

    thin_border = openpyxl.styles.Border(
        bottom=openpyxl.styles.Side(style="thin", color="FFB0B0B0"),
        right=openpyxl.styles.Side(style="thin", color="FFB0B0B0"),
    )
    for r in range(1, row + 1):
        for c in range(1, HARDWARE_END_COL + 1):
            ws.cell(row=r, column=c).border = thin_border

    ws.freeze_panes = "A2"
    return row + 1


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__.strip())
        return 1

    wb_path = argv[1]
    threshold_days = int(argv[2]) if len(argv) > 2 else 42

    success, message, results, hardware = analyse_workbook(wb_path, threshold_days)
    print(f"{message}\n")

    if success:
        print_summary(results, threshold_days)
        success, excel_msg = write_summary_excel(results, threshold_days, hardware=hardware)
        print(excel_msg)
    return 0 if success else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
