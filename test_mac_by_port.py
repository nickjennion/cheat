# test_mac_by_port.py

MIXED_OUTPUTS = {
    "acc1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 10     aaaa.aaaa.aaaa    DYNAMIC     Gi1/0/1",
        " 20     bbbb.bbbb.bbbb    DYNAMIC     Gi1/0/24",
        "Total Mac Addresses for this criterion: 2",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Device ID: SEP001122334455",
        "Entry address(es):",
        "  IP address: 10.0.0.2",
        "Platform: Cisco IP Phone 8845,  Capabilities: Host",
        "Interface: GigabitEthernet1/0/1,  Port ID (outgoing port): GigabitEthernet1/0/1",
        "-------------------------",
        "Device ID: dist1",
        "Entry address(es):",
        "  IP address: 10.0.0.3",
        "Platform: cisco WS-C3850-48P,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/0/24,  Port ID (outgoing port): GigabitEthernet1/0/24",
        "Total cdp entries displayed : 2",
    ]),
    "dist1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 20     cccc.cccc.cccc    DYNAMIC     Te1/1/1",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Device ID: acc1",
        "Entry address(es):",
        "  IP address: 10.0.0.1",
        "Platform: cisco WS-C3560X-24,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/0/24,  Port ID (outgoing port): GigabitEthernet1/0/24",
        "Total cdp entries displayed : 1",
    ]),
}


def test_build_mac_by_port_report_keeps_all_vlans_and_ports():
    from mac_by_port import build_mac_by_port_report
    report = build_mac_by_port_report(MIXED_OUTPUTS)
    assert sorted(r.vlan for r in report.rows) == ["10", "20", "20"]
    assert len(report.rows) == 3


def test_build_mac_by_port_report_keeps_child_switch_uplink():
    from mac_by_port import build_mac_by_port_report
    report = build_mac_by_port_report(MIXED_OUTPUTS)
    row = next(r for r in report.rows
               if r.switch == "acc1" and r.interface == "Gi1/0/24")
    assert row.device_type == "Switch/router"
    assert row.neighbour == "dist1 (Gi1/0/24)"
    assert report.uplink_rows == 1


def test_build_mac_by_port_report_labels_ip_phone():
    from mac_by_port import build_mac_by_port_report
    report = build_mac_by_port_report(MIXED_OUTPUTS)
    row = next(r for r in report.rows
               if r.switch == "acc1" and r.interface == "Gi1/0/1")
    assert row.device_type == "IP phone"
    assert row.neighbour.startswith("SEP001122334455")


def test_build_mac_by_port_report_summary_and_counts():
    from mac_by_port import build_mac_by_port_report
    report = build_mac_by_port_report(MIXED_OUTPUTS)
    assert report.vlan_stacks == {"10": ["acc1"], "20": ["acc1", "dist1"]}
    assert report.vlan_counts == {"10": 1, "20": 2}


MULTI_MAC_OUTPUTS = {
    "acc1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 10     aaaa.aaaa.aaaa    DYNAMIC     Gi1/0/5",
        " 10     bbbb.bbbb.bbbb    DYNAMIC     Gi1/0/5",
        "Total Mac Addresses for this criterion: 2",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Total cdp entries displayed : 0",
    ]),
}


def test_build_mac_by_port_report_flags_multi_mac_port():
    from mac_by_port import build_mac_by_port_report
    report = build_mac_by_port_report(MULTI_MAC_OUTPUTS)
    assert len(report.rows) == 2
    assert all("Multiple MACs" in r.notes for r in report.rows)


def test_write_mac_by_port_report_excel_and_csv(tmp_path):
    import csv
    import openpyxl
    from mac_by_port import build_mac_by_port_report
    from excel_generator import (
        write_mac_by_port_report_excel,
        write_mac_by_port_report_csv,
        MAC_BY_PORT_HEADERS,
    )

    report = build_mac_by_port_report(MIXED_OUTPUTS)

    xlsx = tmp_path / "byport.xlsx"
    ok, _ = write_mac_by_port_report_excel(report, str(xlsx))
    assert ok
    ws = openpyxl.load_workbook(xlsx)["MACs By Port"]
    assert ws.cell(row=1, column=1).value == "MACs Found On These Switches (All VLANs)"
    hdr_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == "Switch")
    headers = [ws.cell(row=hdr_row, column=c).value for c in range(1, len(MAC_BY_PORT_HEADERS) + 1)]
    assert headers == MAC_BY_PORT_HEADERS

    cpath = tmp_path / "byport.csv"
    ok, _ = write_mac_by_port_report_csv(report, str(cpath))
    assert ok
    with open(cpath, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == MAC_BY_PORT_HEADERS
    assert len(rows) == 1 + len(report.rows)
