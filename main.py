#!/usr/bin/env python3
import json
import sys
import getpass
import re
import time
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import Optional
from dnac_client import DNACClient

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Data Models
# ============================================================================

@dataclass
class StackMember:
    member_num: int = 0
    model: str = ""
    uptime: str = ""
    sw_version: str = ""
    is_active: bool = False


@dataclass
class InterfaceRecord:
    switch: str = ""
    stack_member: str = ""
    model: str = ""
    uptime: str = ""
    sw_version: str = ""
    iface: str = ""
    description: str = ""
    state: str = ""
    protocol: str = ""
    last_input: str = ""
    vlan: str = ""
    counters_in: str = ""
    suspect: str = ""


# ============================================================================
# Hardcoded Commands for Command Runner
# ============================================================================

DNAC_COMMANDS = [
    "show hardware",
    "show interfaces",
    "show interfaces status",
    "show interface counters"
]


# ============================================================================
# Regex Patterns (Interface Parser)
# ============================================================================

RE_IFACE_HEADER = re.compile(
    r'^((?:GigabitEthernet|TenGigabitEthernet|FastEthernet|'
    r'FortyGigabitEthernet|HundredGigE)\S+)\s+is\s+(\S+(?:\s+\S+)?),\s*'
    r'line protocol is\s+(\S+)',
    re.IGNORECASE
)
RE_DESCRIPTION = re.compile(r'^\s+Description:\s+(.*)', re.IGNORECASE)
RE_LAST_INPUT = re.compile(r'^\s+Last input\s+(\S+),', re.IGNORECASE)
RE_STATUS_HEADER = re.compile(r'^Port\s+Name\s+Status\s+Vlan', re.IGNORECASE)
RE_STATUS_PORT = re.compile(r'^((?:Gi|Te|Fa|Fo|Hu|Po)\d+(?:/\d+){1,3})\s+', re.IGNORECASE)
RE_STATUS_FIND = re.compile(
    r'\b(connected|notconnect|disabled|err-disabled|inactive|sfpAbsent|xcvrAbsent)\s+(\S+)',
    re.IGNORECASE
)
RE_COUNTERS_HEADER = re.compile(r'^Port\s+InOctets', re.IGNORECASE)
RE_COUNTERS_ROW = re.compile(
    r'^((?:Gi|Te|Fa|Fo|Hu)\d+/\d+/\d+(?:/\d+)?)\s+(\d+)',
    re.IGNORECASE
)
RE_STACK_TABLE_ROW = re.compile(
    r'^(\*?)\s+(\d+)\s+\d+\s+'
    r'(WS-\S+|C\d+\S*|\S+-\S+)\s+'
    r'(\S+)\s+',
    re.IGNORECASE
)
RE_SWITCH_SECTION = re.compile(r'^Switch\s+0*(\d+)\s*$', re.IGNORECASE)
RE_SWITCH_UPTIME = re.compile(r'Switch\s+uptime\s*:\s*(.+)', re.IGNORECASE)
RE_HOSTNAME_UPTIME = re.compile(r'^\S+\s+uptime\s+is\s+(.+)', re.IGNORECASE)
RE_MODEL_NUMBER = re.compile(r'Model Number\s*:\s*(\S+)', re.IGNORECASE)
RE_SHOW_HW_TRIGGER = re.compile(r'show\s+(hardware|version)', re.IGNORECASE)


# ============================================================================
# Interface Parser Functions
# ============================================================================

def shorten_iface(name: str) -> str:
    for long, short in [
        ("TenGigabitEthernet", "Te"),
        ("GigabitEthernet", "Gi"),
        ("FastEthernet", "Fa"),
        ("FortyGigabitEthernet", "Fo"),
        ("HundredGigE", "Hu"),
    ]:
        if name.lower().startswith(long.lower()):
            return short + name[len(long):]
    return name


def member_from_iface(iface: str) -> str:
    m = re.match(r'[A-Za-z]+(\d+)/', iface)
    return m.group(1) if m else ""


def parse_status_row(line):
    m = RE_STATUS_PORT.match(line)
    if not m:
        return None
    port = m.group(1)
    m2 = RE_STATUS_FIND.search(line)
    if not m2:
        return None
    return port, m2.group(1).lower(), m2.group(2)


def parse_hardware(lines: list[str]) -> dict[int, StackMember]:
    members: dict[int, StackMember] = {}
    in_hw = False
    in_stack_table = False
    current_member: Optional[int] = None

    for line in lines:
        s = line.rstrip()

        if RE_SHOW_HW_TRIGGER.search(s):
            in_hw = True
            in_stack_table = False
            current_member = None
            continue

        if not in_hw:
            continue

        m = RE_SWITCH_SECTION.match(s)
        if m:
            current_member = int(m.group(1))
            in_stack_table = False
            if current_member not in members:
                members[current_member] = StackMember(member_num=current_member)
            continue

        if current_member is not None:
            if re.match(r'^-{4,}', s):
                continue

            m = RE_SWITCH_UPTIME.search(s)
            if m:
                members[current_member].uptime = m.group(1).strip()
                continue

            m = RE_MODEL_NUMBER.search(s)
            if m:
                if not members[current_member].model:
                    members[current_member].model = m.group(1).strip()
                continue

            if RE_SHOW_HW_TRIGGER.search(s):
                current_member = None
                in_hw = False
            continue

        if re.match(r'^-{4,}', s) and in_hw:
            in_stack_table = True
            continue

        m = RE_HOSTNAME_UPTIME.match(s)
        if m and in_hw and not in_stack_table:
            if 1 not in members:
                members[1] = StackMember(member_num=1)
            if not members[1].uptime:
                members[1].uptime = m.group(1).strip()
            continue

        if in_stack_table:
            m = RE_STACK_TABLE_ROW.match(s)
            if m:
                is_active = m.group(1) == "*"
                member_num = int(m.group(2))
                model = m.group(3)
                sw_ver = m.group(4)
                if member_num not in members:
                    members[member_num] = StackMember(member_num=member_num)
                members[member_num].model = model
                members[member_num].sw_version = sw_ver
                members[member_num].is_active = is_active
                continue
            if not s.strip():
                in_stack_table = False
            continue

    return members


def parse_output(text: str, hostname: str) -> tuple[list[InterfaceRecord], dict[int, StackMember]]:
    lines = text.split('\n')
    stack_members = parse_hardware(lines)

    int_data: dict[str, InterfaceRecord] = {}
    current_iface: Optional[str] = None
    in_show_interfaces = False
    in_show_status = False
    in_show_counters = False

    for line in lines:
        s = line.rstrip()

        if RE_STATUS_HEADER.match(s):
            in_show_status = True
            in_show_interfaces = False
            in_show_counters = False
            current_iface = None
            continue

        if RE_COUNTERS_HEADER.match(s):
            in_show_counters = True
            in_show_status = False
            in_show_interfaces = False
            current_iface = None
            continue

        m = RE_IFACE_HEADER.match(s)
        if m:
            in_show_interfaces = True
            in_show_status = False
            in_show_counters = False

            full_name = m.group(1)
            hw_state = m.group(2).strip()
            proto = m.group(3).strip()
            short = shorten_iface(full_name)
            current_iface = short

            if short not in int_data:
                int_data[short] = InterfaceRecord()

            rec = int_data[short]
            rec.switch = hostname
            rec.iface = short

            if "administratively" in hw_state.lower():
                rec.state = "disabled"
            elif hw_state.lower() == "up":
                rec.state = "connected"
            else:
                rec.state = hw_state.lower()

            rec.protocol = proto.lower().split()[0]
            continue

        if in_show_interfaces and current_iface:
            m = RE_DESCRIPTION.match(s)
            if m:
                int_data[current_iface].description = m.group(1).strip()
                continue

            m = RE_LAST_INPUT.match(s)
            if m:
                int_data[current_iface].last_input = m.group(1).strip()
                continue

        if in_show_status:
            result = parse_status_row(s)
            if result:
                short, status, vlan = result
                if short not in int_data:
                    int_data[short] = InterfaceRecord(switch=hostname, iface=short)
                int_data[short].state = status
                int_data[short].vlan = vlan
            continue

        if in_show_counters:
            m = RE_COUNTERS_ROW.match(s)
            if m:
                short = m.group(1)
                if short not in int_data:
                    int_data[short] = InterfaceRecord(switch=hostname, iface=short)
                int_data[short].counters_in = m.group(2)
            continue

    for rec in int_data.values():
        member_str = member_from_iface(rec.iface)
        rec.stack_member = member_str

        if member_str and stack_members:
            try:
                mn = int(member_str)
                if mn in stack_members:
                    sm = stack_members[mn]
                    rec.model = sm.model
                    rec.uptime = sm.uptime
                    rec.sw_version = sm.sw_version
            except ValueError:
                pass

        rec.suspect = "NO" if (not rec.last_input or rec.last_input.lower() == "never") else "YES"

    def sort_key(rec):
        parts = re.findall(r'\d+', rec.iface)
        prefix = re.match(r'[A-Za-z]+', rec.iface)
        return (prefix.group() if prefix else "", [int(x) for x in parts])

    return sorted(int_data.values(), key=sort_key), stack_members


def uptime_days(uptime_str: str) -> Optional[float]:
    if not uptime_str:
        return None
    total = 0.0
    for val, unit in re.findall(r'(\d+)\s+(week|day|hour|minute)', uptime_str, re.I):
        v = int(val)
        if "week" in unit:
            total += v * 7
        elif "day" in unit:
            total += v
        elif "hour" in unit:
            total += v / 24
    return total if total > 0 else None


# ============================================================================
# Excel Generation
# ============================================================================

HEADERS = [
    "Switch",
    "Stack Member",
    "Model",
    "SW Version",
    "Member Uptime (days)",
    "Interface",
    "Description",
    "State",
    "Protocol",
    "VLAN",
    "Counters In (Octets)",
    "Last Input",
    "Suspect (Has Had Traffic)",
]

COL_WIDTHS = [28, 13, 18, 12, 20, 12, 36, 14, 10, 8, 22, 14, 26]

STATE_COLOURS = {
    "connected": "FFD4EDDA",
    "notconnect": "FFFFF3CD",
    "disabled": "FFE2E3E5",
    "err-disabled": "FFF8D7DA",
}

SUSPECT_COLOUR = "FFFFD700"
UPTIME_COLOUR = "FFFCE4D6"


def write_excel_sheet(ws, records: list[InterfaceRecord], stack_members: dict[int, StackMember]):
    header_font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="FF2B579A")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="FFB0B0B0"),
        right=Side(style="thin", color="FFB0B0B0"),
    )

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center")

    UPTIME_THRESHOLD_DAYS = 42

    for row_idx, rec in enumerate(records, start=2):
        days = uptime_days(rec.uptime)
        values = [
            rec.switch,
            rec.stack_member,
            rec.model,
            rec.sw_version,
            round(days, 1) if days is not None else "",
            rec.iface,
            rec.description,
            rec.state,
            rec.protocol,
            rec.vlan,
            rec.counters_in,
            rec.last_input,
            rec.suspect,
        ]

        state_key = rec.state.lower() if rec.state else ""
        row_colour = STATE_COLOURS.get(state_key, "FFFFFFFF")
        short_up = days is not None and days < UPTIME_THRESHOLD_DAYS

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            if col == 13 and value == "YES":
                cell.fill = PatternFill("solid", start_color=SUSPECT_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            elif col == 5 and short_up:
                cell.fill = PatternFill("solid", start_color=UPTIME_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=row_colour)

    ws.auto_filter.ref = ws.dimensions


def write_excel(devices_data: dict[str, tuple[list[InterfaceRecord], dict[int, StackMember]]], outpath: str):
    """Write Excel with one sheet per device."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for hostname, (records, stack_members) in devices_data.items():
        ws = wb.create_sheet(title=hostname[:31])
        write_excel_sheet(ws, records, stack_members)

    wb.save(outpath)
    total_records = sum(len(records) for records, _ in devices_data.values())
    print(f"✓ Saved: {outpath}")


# ============================================================================
# CLI Functions
# ============================================================================

def get_credentials() -> tuple[str, str, str]:
    print("=" * 60)
    print("Cisco DNAC Device Query Tool")
    print("=" * 60)
    print()

    host = input("Enter DNAC server hostname/IP: ").strip()
    if not host:
        print("Error: hostname/IP is required")
        sys.exit(1)

    username = input("Enter username: ").strip()
    if not username:
        print("Error: username is required")
        sys.exit(1)

    password = getpass.getpass("Enter password: ")
    if not password:
        print("Error: password is required")
        sys.exit(1)

    return host, username, password


def authenticate_and_fetch(host: str, username: str, password: str) -> tuple[list[dict], object] | None:
    print("\nAuthenticating...")
    client = DNACClient(host, username, password)

    if not client.authenticate():
        print("Error: Failed to authenticate with DNAC")
        return None

    print("✓ Authentication successful")
    print("\nFetching devices...")
    devices = client.get_devices()

    if not devices:
        print("No devices found or failed to retrieve devices")
        return None

    print(f"✓ Found {len(devices)} devices")
    return devices, client


def save_devices(devices: list[dict], filename: str = "all_devices.json") -> bool:
    try:
        with open(filename, "w") as f:
            json.dump(devices, f, indent=2)
        print(f"✓ Saved {len(devices)} devices to {filename}")
        return True
    except IOError as e:
        print(f"Error saving devices: {e}")
        return False


def filter_devices_by_hostname(devices: list[dict], hostname_filter: str) -> list[dict]:
    pattern = hostname_filter.lower()
    return [d for d in devices if pattern in d.get("hostname", "").lower()]


def display_devices(devices: list[dict]) -> None:
    if not devices:
        print("No devices found.")
        return

    print(f"\n{'#':<3} {'Hostname':<25} {'Model':<20} {'IP Address':<15} {'Serial':<20} {'UUID':<36}")
    print("-" * 119)

    for idx, device in enumerate(devices, start=1):
        hostname = device.get("hostname", "N/A")[:24]
        model = device.get("type", "N/A")[:19]
        ip = device.get("managementIpAddress", "N/A")[:14]
        serial = device.get("serialNumber", "N/A")[:19]
        uuid = device.get("id", "N/A")[:35]
        print(f"{idx:<3} {hostname:<25} {model:<20} {ip:<15} {serial:<20} {uuid:<36}")


def select_devices(devices: list[dict]) -> list[dict]:
    """Interactive device selection (single or batch)."""
    while True:
        print("\nOptions:")
        print("  's' - Select single device")
        print("  'b' - Select batch of devices")
        print("  'f' - Filter and try again")
        print("  'q' - Quit")
        choice = input("\nChoice: ").strip().lower()

        if choice == "q":
            return []

        if choice == "f":
            return None

        if choice == "s":
            device_num = input("Enter device number: ").strip()
            try:
                idx = int(device_num) - 1
                if 0 <= idx < len(devices):
                    return [devices[idx]]
                else:
                    print("Invalid device number")
            except ValueError:
                print("Please enter a number")
            continue

        if choice == "b":
            device_nums = input("Enter device numbers (comma-separated): ").strip()
            selected = []
            try:
                for num_str in device_nums.split(","):
                    idx = int(num_str.strip()) - 1
                    if 0 <= idx < len(devices):
                        selected.append(devices[idx])
                    else:
                        print(f"Skipped invalid device number: {num_str.strip()}")
                if selected:
                    return selected
                else:
                    print("No valid devices selected")
            except ValueError:
                print("Please enter numbers separated by commas")
            continue


def execute_on_devices(selected_devices: list[dict], client: DNACClient, session_timestamp: str) -> dict[str, str]:
    """Execute commands on selected devices and save outputs."""
    outputs = {}

    for device in selected_devices:
        hostname = device.get("hostname", "unknown")
        device_id = device.get("id")

        print(f"\n{'='*60}")
        print(f"Executing commands on: {hostname}")
        print(f"{'='*60}")

        task_id = client.execute_commands(device_id, DNAC_COMMANDS)
        if not task_id:
            print(f"✗ Failed to start command execution on {hostname}")
            continue

        print(f"Task ID: {task_id}")
        print("Polling for results (30s timeout)...")

        result = None
        for i in range(30):
            time.sleep(1)
            task_result = client.get_task_result(task_id)
            if task_result and task_result.get("endTime"):
                result = task_result
                break
            print(f"  [{i+1}/30] Waiting...")

        if not result:
            print(f"✗ Command execution timed out on {hostname}")
            continue

        output_text = result.get("result", "")
        if not output_text:
            print(f"✗ No output received from {hostname}")
            continue

        filename = f"command_output_{hostname}_{session_timestamp}.txt"
        try:
            with open(filename, "w") as f:
                f.write(output_text)
            outputs[hostname] = output_text
            print(f"✓ Output saved to {filename}")
        except IOError as e:
            print(f"✗ Failed to save output: {e}")

    return outputs


def parse_and_generate_excel(outputs: dict[str, str], session_timestamp: str) -> bool:
    """Parse command outputs and generate Excel."""
    if not outputs:
        print("No command outputs to parse")
        return False

    print("\n" + "=" * 60)
    print("Parsing command outputs and generating Excel...")
    print("=" * 60)

    devices_data = {}
    for hostname, output_text in outputs.items():
        print(f"\nParsing {hostname}...", end=" ")
        records, stack_members = parse_output(output_text, hostname)
        if records:
            devices_data[hostname] = (records, stack_members)
            print(f"✓ {len(records)} interfaces found")
        else:
            print("✗ No interfaces parsed")

    if not devices_data:
        print("No data to write to Excel")
        return False

    excel_filename = f"unpatching_list_{session_timestamp}.xlsx"
    write_excel(devices_data, excel_filename)
    return True


def main():
    try:
        host, username, password = get_credentials()

        result = authenticate_and_fetch(host, username, password)
        if result is None:
            sys.exit(1)

        devices, client = result
        save_devices(devices)

        session_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        while True:
            print("\n" + "=" * 60)
            hostname_filter = input("Enter hostname filter (or 'quit'): ").strip()

            if hostname_filter.lower() == "quit":
                print("Exiting...")
                sys.exit(0)

            if not hostname_filter:
                print("Please enter a hostname filter")
                continue

            filtered = filter_devices_by_hostname(devices, hostname_filter)
            print(f"\nFound {len(filtered)} device(s):")
            display_devices(filtered)

            selected = select_devices(filtered)
            if selected is None:
                continue
            if not selected:
                continue

            outputs = execute_on_devices(selected, client, session_timestamp)
            parse_and_generate_excel(outputs, session_timestamp)

    except KeyboardInterrupt:
        print("\n\nExiting...")
        sys.exit(0)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
