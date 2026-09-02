import openpyxl

from interface_parser import InterfaceRecord


RAW = {
    "stack-a": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        " 10     aaaa.aaaa.aaaa    DYNAMIC     Gi1/0/1",
        " 20     bbbb.bbbb.bbbb    DYNAMIC     Gi1/0/2",
        "show cdp neighbors detail",
        "Total cdp entries displayed : 0",
        "show ip device tracking all",
        "10.10.10.5     aaaa.aaaa.aaaa  10  GigabitEthernet1/0/1  ACTIVE",
        "10.20.20.5     cccc.cccc.cccc  20  GigabitEthernet1/0/2  INACTIVE",
    ])
}


def _devices():
    return {
        "stack-a": ([
            InterfaceRecord(switch="stack-a", stack_member="1", iface="Gi1/0/1",
                            description="Desk 1", state="connected", vlan="10"),
            InterfaceRecord(switch="stack-a", stack_member="1", iface="Gi1/0/2",
                            description="Desk 2", state="connected", vlan="20"),
            InterfaceRecord(switch="stack-a", stack_member="1", iface="Gi1/0/3",
                            description="Empty", state="notconnect", vlan="30"),
        ], {})
    }


def test_parse_legacy_ip_tracking_filtered_rows():
    from palantir_report import parse_legacy_ip_tracking
    rows = parse_legacy_ip_tracking(RAW["stack-a"])
    assert [(r.ip, r.mac, r.vlan, r.interface, r.state) for r in rows] == [
        ("10.10.10.5", "aaaa.aaaa.aaaa", "10", "Gi1/0/1", "ACTIVE"),
        ("10.20.20.5", "cccc.cccc.cccc", "20", "Gi1/0/2", "INACTIVE"),
    ]


def test_parse_legacy_ip_tracking_standalone_port_and_mac_formats():
    from palantir_report import parse_legacy_ip_tracking

    text = "\n".join([
        "10.50.1.3 00:11:22:33:44:55 501 GigabitEthernet0/3 30 ACTIVE",
        "10.50.1.9 00-22-33-44-55-66 501 Gi0/9 INACTIVE",
        "10.50.1.10 003344556677 501 FastEthernet0/10 REACHABLE",
    ])
    rows = parse_legacy_ip_tracking(text)
    assert [(row.mac, row.interface) for row in rows] == [
        ("0011.2233.4455", "Gi0/3"),
        ("0022.3344.5566", "Gi0/9"),
        ("0033.4455.6677", "Fa0/10"),
    ]


def test_parse_ip_tracking_accepts_modern_wrapped_sisf_rows_and_deduplicates():
    from palantir_report import parse_ip_tracking

    text = "\n".join([
        "ARP 10.10.10.5 aaaa.aaaa.aaaa GigabitEthernet1/0/1 10",
        "10.10.10.5 aaaa.aaaa.aaaa 10 Gi1/0/1 ACTIVE",
        "L 10.10.10.1 0000.0c9f.f001 Vl10 10 0100 5mn REACHABLE",
    ])
    rows = parse_ip_tracking(text)
    assert rows == [
        __import__("palantir_report").LegacyIpTrackingEntry(
            "10.10.10.5", "aaaa.aaaa.aaaa", "10", "Gi1/0/1", "ACTIVE"
        )
    ]


def test_palantir_reports_disabled_tracking_separately():
    from palantir_report import build_palantir_report

    raw = {"stack-a": RAW["stack-a"].split("10.10.10.5", 1)[0]
           + "IP Device Tracking = Disabled\n"}
    report = build_palantir_report(_devices(), raw)
    assert any("disabled on: stack-a" in note for note in report.notes)
    assert not any("No usable" in note for note in report.notes)


def test_palantir_reports_command_failure_separately():
    from palantir_report import build_palantir_report

    raw = {"stack-a": (
        "[Command Runner FAILURE: show ip device tracking all]\n"
        "Command is not supported"
    )}
    report = build_palantir_report(_devices(), raw)
    assert any("command failed on: stack-a" in note for note in report.notes)
    assert not any("No usable" in note for note in report.notes)


def test_palantir_correlates_and_retains_empty_ports():
    from palantir_report import build_palantir_report
    report = build_palantir_report(_devices(), RAW)

    matched = next(r for r in report.rows if r.mac == "aaaa.aaaa.aaaa")
    assert matched.port.iface == "Gi1/0/1"
    assert matched.client_ip == "10.10.10.5"
    assert matched.client_vlan == "10"

    mac_only = next(r for r in report.rows if r.mac == "bbbb.bbbb.bbbb")
    assert mac_only.client_ip == ""
    assert report.mac_rows_without_ip == 1

    tracked_only = next(r for r in report.rows if r.mac == "cccc.cccc.cccc")
    assert tracked_only.client_ip == "10.20.20.5"
    assert "not present in MAC table" in tracked_only.notes

    empty = [r for r in report.rows if r.port.iface == "Gi1/0/3"]
    assert len(empty) == 1
    assert empty[0].mac == ""


def test_write_palantir_workbook_has_all_ports_and_per_stack(tmp_path):
    from excel_generator import HEADERS, PALANTIR_HEADERS, write_palantir_excel
    from palantir_report import build_palantir_report

    devices = _devices()
    report = build_palantir_report(devices, RAW)
    out = tmp_path / "palantir.xlsx"
    ok, _ = write_palantir_excel(report, devices, 42, str(out))
    assert ok

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == [
        "All Ports", "All MAC Addresses", "Port Utilisation", "VLAN Inventory", "stack-a"
    ]
    ws = wb["All Ports"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, len(HEADERS) + 1)]
    assert headers == HEADERS
    assert ws.max_row == len(devices["stack-a"][0]) + 1

    mac_ws = wb["All MAC Addresses"]
    mac_headers = [
        mac_ws.cell(row=1, column=col).value
        for col in range(1, len(PALANTIR_HEADERS) + 1)
    ]
    assert mac_headers == PALANTIR_HEADERS
    assert "MAC Address" in mac_headers
    assert "Manufacturer" in mac_headers
    assert "Client IP" in mac_headers
    assert "Client VLAN" in mac_headers


def test_palantir_all_ports_does_not_expand_for_downstream_macs(tmp_path):
    from excel_generator import write_palantir_excel
    from palantir_report import build_palantir_report

    devices = {"core": (list(_devices()["stack-a"][0]), {})}
    raw = {"core": "\n".join([
        "show mac address-table",
        " 10 aaaa.aaaa.aaaa DYNAMIC Gi1/0/1",
        " 10 bbbb.bbbb.bbbb DYNAMIC Gi1/0/1",
        " 10 cccc.cccc.cccc DYNAMIC Gi1/0/1",
    ])}
    report = build_palantir_report(devices, raw)
    assert len(report.rows) == 5

    out = tmp_path / "palantir-expanded.xlsx"
    ok, _ = write_palantir_excel(report, devices, 42, str(out))
    assert ok
    wb = openpyxl.load_workbook(out)
    assert wb["All Ports"].max_row == 4
    assert wb["All MAC Addresses"].max_row == 6


def test_menu_5_wires_palantir_mode():
    from cheat_constants import build_palantir_command_list

    src = open("main.py", encoding="utf-8").read()
    assert "x) Palantir Mode" in src
    assert 'elif choice == "x"' in src
    commands = build_palantir_command_list(False)
    assert commands.count("show cdp neighbors detail") == 1
    assert "show mac address-table" in commands
    assert "show device-tracking database" in commands
    assert "show ip device tracking all" in commands
    assert all("| include" not in command for command in commands)
    assert "generate_cdp_topology" in src
    assert 'layout="pyramid"' in src
