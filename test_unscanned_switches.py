CDP_BRIEF = "\n".join([
    "show cdp neighbors",
    "Capability Codes: R - Router, T - Trans Bridge, B - Source Route Bridge",
    "                  S - Switch, H - Host, I - IGMP, r - Repeater, P - Phone, ",
    "                  D - Remote, C - CVTA, M - Two-port Mac Relay ",
    "",
    "Device ID        Local Intrfce     Holdtme    Capability  Platform  Port ID",
    "sw4              Gig 0/0           137              S I   C9KV-UADP Gig 0/0",
    "sw2              Gig 1/0/3         169              S I   C9KV-UADP Gig 1/0/2",
    "deskphone-01     Gig 1/0/5         120              H P   IP-Phone  Port 1",
    "",
    "Total cdp entries displayed : 3",
    "sw1#",
])


def test_parse_cdp_switch_neighbors_keeps_only_switches():
    from unscanned_switches import parse_cdp_switch_neighbors
    nbrs = parse_cdp_switch_neighbors(CDP_BRIEF)
    devices = sorted(n.device for n in nbrs)
    assert devices == ["sw2", "sw4"]  # phone excluded


def test_parse_cdp_switch_neighbors_extracts_fields():
    from unscanned_switches import parse_cdp_switch_neighbors
    nbrs = {n.device: n for n in parse_cdp_switch_neighbors(CDP_BRIEF)}
    n = nbrs["sw4"]
    assert n.local_iface == "Gi0/0"
    assert n.capability == "S I"
    assert n.platform == "C9KV-UADP"
    assert n.neighbour_port == "Gi0/0"
    assert n.seen_on == ""


def _sw1_text():
    return "\n".join([
        "Device ID        Local Intrfce     Holdtme    Capability  Platform  Port ID",
        "sw2              Gig 1/0/3         169              S I   C9KV-UADP Gig 1/0/2",
        "sw3              Gig 1/0/1         156              S I   C9KV-UADP Gig 1/0/2",
        "SW4.example.net  Gig 0/0           137              S I   C9KV-UADP Gig 0/0",
        "Total cdp entries displayed : 3",
    ])


def test_find_unscanned_switches_flags_only_unknown():
    from unscanned_switches import find_unscanned_switches
    raw = {"sw1": _sw1_text(), "sw2": "", "sw3": ""}
    rows = find_unscanned_switches(raw, raw.keys())
    # sw2/sw3 are scanned -> excluded; sw4 (FQDN, different case) -> flagged.
    assert [r.device for r in rows] == ["SW4.example.net"]
    assert rows[0].seen_on == "sw1"
    assert rows[0].local_iface == "Gi0/0"


def test_find_unscanned_switches_dedupes_sightings():
    from unscanned_switches import find_unscanned_switches
    raw = {"sw1": _sw1_text(), "sw1b": _sw1_text()}  # same sighting text twice, diff hosts
    rows = find_unscanned_switches(raw, ["sw1", "sw1b", "sw2", "sw3"])
    # sw2/sw3 scanned-out; sw4 seen on two different hosts -> two sightings.
    assert sorted((r.device, r.seen_on) for r in rows) == [
        ("SW4.example.net", "sw1"), ("SW4.example.net", "sw1b")
    ]


def test_write_unscanned_switches_block_with_rows():
    import openpyxl
    from unscanned_switches import SwitchNeighbour
    from excel_generator import write_unscanned_switches_block, UNSCANNED_HEADERS
    ws = openpyxl.Workbook().active
    rows = [SwitchNeighbour("sw4", "C9KV-UADP", "S I", "Gi0/1", "Gi0/2", "sw1")]
    write_unscanned_switches_block(ws, 5, rows)
    assert "Unscanned Cisco Switches" in ws.cell(row=5, column=1).value
    assert [ws.cell(row=6, column=c).value for c in range(1, 7)] == UNSCANNED_HEADERS
    assert ws.cell(row=7, column=1).value == "sw4"
    assert ws.cell(row=7, column=4).value == "sw1"
    assert ws.cell(row=7, column=5).value == "Gi0/1"
    assert ws.cell(row=7, column=6).value == "Gi0/2"


def test_write_unscanned_switches_block_empty():
    import openpyxl
    from excel_generator import write_unscanned_switches_block
    ws = openpyxl.Workbook().active
    write_unscanned_switches_block(ws, 5, [])
    assert ws.cell(row=6, column=1).value == "None detected"


def _devices_data():
    from interface_parser import InterfaceRecord
    rec = InterfaceRecord(switch="sw1", iface="Gi1/0/1", stack_member="1",
                          last_input="00:00:01")
    return {"sw1": ([rec], {})}


def _find_block_title_row(ws):
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and "Unscanned Cisco Switches" in str(val):
            return r
    return None


def test_combined_excel_includes_block_below_total(tmp_path):
    import openpyxl
    from excel_generator import write_combined_excel
    from unscanned_switches import SwitchNeighbour
    out = tmp_path / "report.xlsx"
    unscanned = [SwitchNeighbour("sw4", "C9KV-UADP", "S I", "Gi0/0", "Gi0/0", "sw1")]
    ok, _ = write_combined_excel(_devices_data(), 42, str(out), unscanned=unscanned)
    assert ok
    ws = openpyxl.load_workbook(out)["Port Utilisation"]
    title_row = _find_block_title_row(ws)
    assert title_row is not None
    # Block sits below the TOTAL row of the utilisation table.
    total_row = next(r for r in range(1, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value == "TOTAL")
    assert title_row > total_row
    assert ws.cell(row=title_row + 2, column=1).value == "sw4"


def test_combined_excel_omits_block_when_unscanned_none(tmp_path):
    import openpyxl
    from excel_generator import write_combined_excel
    out = tmp_path / "report.xlsx"
    ok, _ = write_combined_excel(_devices_data(), 42, str(out))  # unscanned defaults None
    assert ok
    ws = openpyxl.load_workbook(out)["Port Utilisation"]
    assert _find_block_title_row(ws) is None


def test_generate_excel_mode3_writes_block(tmp_path, monkeypatch):
    import openpyxl
    import cheat_core
    monkeypatch.chdir(tmp_path)
    raw = {"sw1": _sw1_text(), "sw2": "", "sw3": ""}
    results = cheat_core.generate_excel(
        _devices_data(), 3, "report", threshold=42, raw_outputs=raw
    )
    assert results and results[0][0] is True
    path = results[0][1].split("Saved: ")[1].split(" (")[0]
    ws = openpyxl.load_workbook(path)["Port Utilisation"]
    title_row = _find_block_title_row(ws)
    assert title_row is not None
    # sw4 from _sw1_text() is the only unscanned switch.
    assert ws.cell(row=title_row + 2, column=1).value == "SW4.example.net"
