"""
Consolidate a per-stack port report into a single flat sheet.

The main tool writes one worksheet per device/stack. This standalone utility
reads such a workbook and produces a second workbook containing every port for
every hostname on ONE sheet, preserving the same columns and colour coding.

Usage:
    python consolidate_report.py <input.xlsx> [output.xlsx]

If no output path is given, the consolidated file is written alongside the
input as "<name>-consolidated-YYYY-MM-DD-HH-MM.xlsx".
"""

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_generator import (
    HEADERS,
    COL_WIDTHS,
    STATE_COLOURS,
    SUSPECT_COLOUR,
    UPTIME_COLOUR,
    UPTIME_THRESHOLD_DAYS,
    get_header_styles,
    get_data_styles,
)

# Column positions are resolved by header name so reordering HEADERS in
# excel_generator.py does not silently misplace the colour logic here.
STATE_COL = HEADERS.index("State") + 1
SUSPECT_COL = HEADERS.index("Suspect (Has Had Traffic)") + 1
UPTIME_COL = HEADERS.index("Member Uptime (days)") + 1

SHEET_NAME = "All Ports"


def read_all_rows(wb) -> list[list]:
    """Collect every data row (row 2+) from every worksheet into one list.

    Rows keep their per-sheet order, so output is grouped by hostname/stack.
    Sheets whose header row does not match HEADERS are skipped with a warning.
    """
    all_rows: list[list] = []
    ncols = len(HEADERS)

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        print(f"  Reading sheet {sheet_idx}/{len(wb.worksheets)}: '{ws.title}'...", end=" ", flush=True)

        if ws.max_row < 2:
            print("(empty, skipped)")
            continue

        # Verify header matches expected
        header = [ws.cell(row=1, column=c).value for c in range(1, ncols + 1)]
        if header != HEADERS:
            print(f"✗ unexpected header layout, skipped")
            continue

        # Extract all data rows
        row_count = 0
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row=row, column=c).value for c in range(1, ncols + 1)]
            # Skip fully blank rows that openpyxl sometimes reports at the tail.
            if all(v is None or v == "" for v in values):
                continue
            all_rows.append(values)
            row_count += 1

        print(f"({row_count} rows)")

    return all_rows


def write_consolidated_sheet(ws, rows: list[list]) -> int:
    """Write all rows to a single worksheet, re-applying the report styling."""
    header_font, header_fill, header_align, header_border = get_header_styles()
    data_font, data_align, data_border = get_data_styles()

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, values in enumerate(rows, start=2):
        state = (values[STATE_COL - 1] or "")
        state_key = str(state).lower()
        row_colour = STATE_COLOURS.get(state_key, "FFFFFFFF")

        uptime_val = values[UPTIME_COL - 1]
        short_up = isinstance(uptime_val, (int, float)) and uptime_val < UPTIME_THRESHOLD_DAYS

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border

            if col == SUSPECT_COL and value == "YES":
                cell.fill = PatternFill("solid", start_color=SUSPECT_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            elif col == UPTIME_COL and short_up:
                cell.fill = PatternFill("solid", start_color=UPTIME_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=row_colour)

    ws.auto_filter.ref = ws.dimensions
    return len(rows)


def consolidate(input_path: str, output_path: str | None = None) -> tuple[bool, str]:
    """Read a per-stack workbook and write a single-sheet consolidated workbook."""
    in_path = Path(input_path).resolve()
    if not in_path.is_file():
        return False, f"✗ Input file not found: {input_path}"

    if output_path is None:
        stamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
        excel_dir = Path("excel_reports").resolve()
        if excel_dir.is_dir():
            output_path = str(excel_dir / f"{in_path.stem}-consolidated-{stamp}.xlsx")
        else:
            output_path = str(in_path.parent / f"{in_path.stem}-consolidated-{stamp}.xlsx")

    print(f"Loading '{input_path}'...")
    try:
        src = openpyxl.load_workbook(in_path, data_only=True)
    except Exception as e:
        return False, f"✗ Failed to open '{input_path}': {e}"

    sheet_count = len(src.worksheets)
    print(f"Found {sheet_count} sheet(s). Reading data...")
    try:
        rows = read_all_rows(src)
    except Exception as e:
        src.close()
        return False, f"✗ Failed to read sheets: {e}"
    finally:
        src.close()

    if not rows:
        return False, "✗ No data rows found across any sheet"

    print(f"Consolidating {len(rows)} rows onto single sheet...")
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = SHEET_NAME
    try:
        count = write_consolidated_sheet(ws, rows)
    except Exception as e:
        return False, f"✗ Failed to consolidate: {e}"

    print(f"Writing '{output_path}'...")
    try:
        out.save(output_path)
    except Exception as e:
        return False, f"✗ Failed to write '{output_path}': {e}"

    return True, (
        f"✓ Consolidated {count} ports from {sheet_count} sheet(s) "
        f"into {output_path}"
    )


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__.strip())
        return 1

    input_path = argv[1]
    output_path = argv[2] if len(argv) > 2 else None
    success, message = consolidate(input_path, output_path)
    print(message)
    return 0 if success else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
