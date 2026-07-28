"""
Excel report generation for CHEAT UNPLUGGED.

Generates formatted Excel workbooks from parsed interface data.
One sheet per device with color-coded interface inventory.
"""

from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from cdp_detail import is_access_point
from interface_parser import InterfaceRecord, StackMember, uptime_days, site_location
from port_utilisation import is_copper_port, write_utilisation_sheet
from unscanned_switches import find_unscanned_switches, SwitchNeighbour
from time_utils import parse_duration_days


# ============================================================================
# Excel Configuration
# ============================================================================

HEADERS = [
    "Switch",
    "Site",
    "Location",
    "Stack Member",
    "Model",
    "SW Version",
    "Member Uptime (days)",
    "Interface",
    "Description",
    "State",
    "Protocol",
    "VLAN",
    "Speed",
    "Type",
    "Counters In (Octets)",
    "Last Input",
    "Suspect (Has Had Traffic)",
    "CDP Neighbors",
]

COL_WIDTHS = [28, 10, 12, 13, 18, 12, 20, 12, 36, 14, 10, 8, 10, 12, 22, 14, 26, 30]

UNSCANNED_TITLE = "Discovered Devices w/ Switching Capability (seen via CDP, not in DNAC)"
UNSCANNED_AP_TITLE = "Discovered Access Points"
UNSCANNED_HEADERS = [
    "Unknown Neighbour", "Platform", "Mgmt IP", "Capability",
    "Seen On", "Local Interface", "Neighbour Port",
]

STATE_COLOURS = {
    "connected": "FFD4EDDA",
    "notconnect": "FFFFF3CD",
    "disabled": "FFE2E3E5",
    "err-disabled": "FFF8D7DA",
}

SUSPECT_COLOUR = "FFFFD700"
UPTIME_COLOUR = "FFFCE4D6"

UPTIME_THRESHOLD_DAYS = 42


# ============================================================================
# Styling Helpers
# ============================================================================

def get_header_styles():
    """Return styled components for header row."""
    font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    fill = PatternFill("solid", start_color="FF2B579A")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="thin", color="FFB0B0B0"),
        right=Side(style="thin", color="FFB0B0B0"),
    )
    return font, fill, align, border


def get_data_styles():
    """Return styled components for data rows."""
    font = Font(name="Arial", size=10)
    align = Alignment(vertical="center")
    border = Border(
        bottom=Side(style="thin", color="FFB0B0B0"),
        right=Side(style="thin", color="FFB0B0B0"),
    )
    return font, align, border


# ============================================================================
# Sheet Writing
# ============================================================================

def write_excel_sheet(ws, records: list[InterfaceRecord], stack_members: dict[int, StackMember]) -> int:
    """
    Write parsed data to an Excel worksheet.
    Returns number of records written.
    """
    header_font, header_fill, header_align, header_border = get_header_styles()
    data_font, data_align, data_border = get_data_styles()

    include_link_state = any(r.last_link_change for r in records)
    headers = HEADERS + (["Last Link Change"] if include_link_state else [])
    col_widths = COL_WIDTHS + ([18] if include_link_state else [])

    # Write headers
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, rec in enumerate(records, start=2):
        days = uptime_days(rec.uptime)
        site, location = site_location(rec.switch)
        values = [
            rec.switch,
            site,
            location,
            rec.stack_member,
            rec.model,
            rec.sw_version,
            round(days, 1) if days is not None else "",
            rec.iface,
            rec.description,
            rec.state,
            rec.protocol,
            rec.vlan,
            rec.speed,
            rec.if_type,
            rec.counters_in,
            rec.last_input,
            rec.suspect,
            rec.cdp_neighbors,
        ]

        if include_link_state:
            values = values + [rec.last_link_change]

        state_key = rec.state.lower() if rec.state else ""
        row_colour = STATE_COLOURS.get(state_key, "FFFFFFFF")
        short_up = days is not None and days < UPTIME_THRESHOLD_DAYS

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border

            # Highlight suspect interfaces (gold) — col 17 after Site+Location insertion
            if col == 17 and value == "YES":
                cell.fill = PatternFill("solid", start_color=SUSPECT_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            # Highlight short uptime (orange) — col 7 after Site+Location insertion
            elif col == 7 and short_up:
                cell.fill = PatternFill("solid", start_color=UPTIME_COLOUR)
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=row_colour)

    ws.auto_filter.ref = ws.dimensions
    return len(records)


# ============================================================================
# Workbook Writing
# ============================================================================

def write_excel(
    devices_data: dict[str, tuple[list[InterfaceRecord], dict[int, StackMember]]],
    outpath: str
) -> tuple[bool, str]:
    """
    Write multi-sheet Excel workbook (one sheet per device).
    Returns (success: bool, message: str)
    """
    if not devices_data:
        return False, "No device data to write"

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        total_records = 0
        total_devices = 0

        for hostname, (records, stack_members) in devices_data.items():
            if not records:
                continue

            # Truncate hostname to 31 chars (Excel sheet name limit)
            sheet_name = hostname[:31]

            # Handle duplicate sheet names by appending suffix
            if sheet_name in wb.sheetnames:
                for i in range(2, 100):
                    test_name = f"{sheet_name[:27]}_{i}"
                    if test_name not in wb.sheetnames:
                        sheet_name = test_name
                        break
                else:
                    raise ValueError(
                        f"Cannot create unique sheet name for hostname prefix '{sheet_name[:27]}'"
                    )

            ws = wb.create_sheet(title=sheet_name)
            count = write_excel_sheet(ws, records, stack_members)
            total_records += count
            total_devices += 1

        wb.save(outpath)
        msg = f"✓ Saved: {outpath} ({total_records} interfaces across {total_devices} devices)"
        return True, msg

    except Exception as e:
        return False, f"✗ Failed to write Excel: {e}"


def _compute_utilisation(
    devices_data: dict,
    threshold_days: int
) -> dict[str, tuple[int, int]]:
    """Compute port utilisation from in-memory records. Returns {hostname: (in_use, idle)}."""
    results: dict[str, tuple[int, int]] = {}
    for hostname, (records, _) in devices_data.items():
        in_use = 0
        idle = 0
        for rec in records:
            if not is_copper_port(rec.iface):
                continue
            days = parse_duration_days(str(rec.last_input).strip() if rec.last_input else "")
            if days is not None and days < threshold_days:
                in_use += 1
            else:
                idle += 1
        if in_use + idle > 0:
            results[hostname] = (in_use, idle)
    return results


def _compute_hardware(devices_data: dict) -> dict[str, dict[int, str]]:
    """Map each host to its stack members' models: {hostname: {member_num: model}}."""
    hardware: dict[str, dict[int, str]] = {}
    for hostname, (_, stack_members) in devices_data.items():
        members = {
            sm.member_num: sm.model
            for sm in stack_members.values()
            if sm.model
        }
        if members:
            hardware[hostname] = members
    return hardware


def _write_neighbour_table(ws, start_row: int, title: str, rows: list) -> int:
    """Write one titled neighbour table starting at start_row.

    Writes a bold title at start_row. With no rows, writes 'None detected'
    below it; otherwise a header row then one row per sighting.
    Returns the next free row (one blank row below the table).
    """
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, name="Arial", size=10)

    if not rows:
        ws.cell(row=start_row + 1, column=1, value="None detected")
        return start_row + 3

    header_font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="FF2B579A")
    header_align = Alignment(horizontal="center", vertical="center")

    hdr_row = start_row + 1
    for col, header in enumerate(UNSCANNED_HEADERS, start=1):
        c = ws.cell(row=hdr_row, column=col, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for i, nb in enumerate(rows):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=nb.device)
        ws.cell(row=r, column=2, value=nb.platform)
        ws.cell(row=r, column=3, value=nb.mgmt_ip)
        ws.cell(row=r, column=4, value=nb.capability)
        ws.cell(row=r, column=5, value=nb.seen_on)
        ws.cell(row=r, column=6, value=nb.local_iface)
        ws.cell(row=r, column=7, value=nb.neighbour_port)

    return hdr_row + 1 + len(rows) + 1


def write_unscanned_switches_block(ws, start_row: int, rows: list) -> None:
    """Append the discovered-neighbours blocks to an existing worksheet.

    Splits rows into switching-capable devices (routers/switches) and access
    points, writing each as its own titled table, devices first.
    """
    devices = [r for r in rows if not is_access_point(r)]
    aps = [r for r in rows if is_access_point(r)]

    next_row = _write_neighbour_table(ws, start_row, UNSCANNED_TITLE, devices)
    _write_neighbour_table(ws, next_row, UNSCANNED_AP_TITLE, aps)


def write_combined_excel(
    devices_data: dict[str, tuple[list[InterfaceRecord], dict[int, StackMember]]],
    threshold_days: int,
    outpath: str,
    unscanned: list | None = None,
) -> tuple[bool, str]:
    """Write single combined workbook: All Ports -> Port Utilisation -> per-stack tabs.

    Sheet order:
      Sheet 1 "All Ports"         -- every port from every device on one sheet
      Sheet 2 "Port Utilisation"  -- copper-port utilisation summary
      Sheets 3-N <hostname>       -- one tab per device/stack
    """
    if not devices_data:
        return False, "No device data to write"

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet 1: All Ports — consolidate all records
        all_records = [rec for records, _ in devices_data.values() for rec in records]
        ws_all = wb.create_sheet(title="All Ports")
        write_excel_sheet(ws_all, all_records, {})

        # Sheet 2: Port Utilisation
        util_results = _compute_utilisation(devices_data, threshold_days)
        util_hardware = _compute_hardware(devices_data)
        ws_util = wb.create_sheet(title="Port Utilisation")
        if util_results:
            next_row = write_utilisation_sheet(
                ws_util, util_results, threshold_days, hardware=util_hardware
            )
        else:
            ws_util.cell(row=1, column=1, value="No copper port data found")
            next_row = 2

        if unscanned is not None:
            write_unscanned_switches_block(ws_util, next_row + 1, unscanned)

        # Sheets 3-N: per-stack
        total_records = len(all_records)
        total_devices = 0
        for hostname, (records, stack_members) in devices_data.items():
            if not records:
                continue
            sheet_name = hostname[:31]
            if sheet_name in wb.sheetnames:
                for i in range(2, 100):
                    candidate = f"{sheet_name[:27]}_{i}"
                    if candidate not in wb.sheetnames:
                        sheet_name = candidate
                        break
                else:
                    raise ValueError(
                        f"Cannot create unique sheet name for hostname prefix '{sheet_name[:27]}'"
                    )
            ws = wb.create_sheet(title=sheet_name)
            write_excel_sheet(ws, records, stack_members)
            total_devices += 1

        wb.save(outpath)
        return True, (
            f"✓ Saved: {outpath} "
            f"({total_records} interfaces across {total_devices} device(s))"
        )

    except Exception as e:
        return False, f"✗ Failed to write Excel: {e}"


# ============================================================================
# Client Search Export
# ============================================================================

CLIENT_SEARCH_HEADERS = [
    "MAC Address",
    "Client Name",
    "Switch",
    "Port",
    "VLAN",
    "IP Address",
    "Status",
    "Type",
    "Vendor",
    "OS Type",
    "Username",
]

CLIENT_SEARCH_COL_WIDTHS = [20, 28, 32, 28, 8, 18, 14, 10, 20, 18, 24]

CLIENT_STATUS_COLOURS = {
    "connected": "FFD4EDDA",
    "inactive": "FFFFF3CD",
    "disconnected": "FFE2E3E5",
}


def write_client_search_excel(clients: list, outpath: str) -> tuple[bool, str]:
    """Write MAC prefix search results to a single-sheet Excel workbook."""
    if not clients:
        return False, "No client records to export"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Client Search"

        header_font, header_fill, header_align, header_border = get_header_styles()
        data_font, data_align, data_border = get_data_styles()

        for col, (header, width) in enumerate(
            zip(CLIENT_SEARCH_HEADERS, CLIENT_SEARCH_COL_WIDTHS), start=1
        ):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        for row_idx, c in enumerate(clients, start=2):
            nd   = c.get("connectedNetworkDevice") or {}
            conn = c.get("connection") or {}
            values = [
                c.get("macAddress") or "",
                c.get("name") or "",
                nd.get("connectedNetworkDeviceName") or "",
                nd.get("interfaceName") or "",
                conn.get("vlanId") or "",
                c.get("ipv4Address") or "",
                c.get("connectionStatus") or "",
                c.get("type") or "",
                c.get("vendor") or "",
                c.get("osType") or "",
                c.get("username") or "",
            ]

            status_key = (c.get("connectionStatus") or "").lower()
            row_colour = CLIENT_STATUS_COLOURS.get(status_key, "FFFFFFFF")
            fill = PatternFill("solid", start_color=row_colour)

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = data_border
                cell.fill = fill

        ws.auto_filter.ref = ws.dimensions
        wb.save(outpath)
        return True, f"✓ Saved: {outpath} ({len(clients)} client(s))"

    except Exception as e:
        return False, f"✗ Failed to write Excel: {e}"
