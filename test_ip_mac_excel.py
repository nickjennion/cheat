"""Excel output for the IP/MAC per VLAN export."""

import openpyxl


def _report(**overrides):
    from ip_mac_report import IpMacReport, IpMacRow
    base = dict(
        vlan_switches={"5": ["sw-01"], "2003": []},
        rows=[
            IpMacRow(switch="sw-01", stack_member="2", interface="Fi2/0/32",
                     vlan="5", ip="10.0.5.10", mac="0011.2233.aaaa",
                     state="REACHABLE", age="145s"),
            IpMacRow(switch="sw-01", stack_member="2", interface="Fi2/0/34",
                     vlan="5", ip="10.0.5.11", mac="0011.2233.bbbb",
                     state="STALE", age="30s",
                     notes="Duplicate IP — held by 2 MACs"),
        ],
    )
    base.update(overrides)
    return IpMacReport(**base)


def _cells(ws, col=1):
    return [ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)]


def test_writes_a_workbook_with_the_expected_sheet(tmp_path):
    from excel_generator import write_ip_mac_report_excel
    out = tmp_path / "ipmac.xlsx"
    ok, msg = write_ip_mac_report_excel(_report(), str(out))
    assert ok is True
    assert "2" in msg
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["IP-MAC Per VLAN"]


def test_summary_block_lists_each_requested_vlan(tmp_path):
    from excel_generator import write_ip_mac_report_excel
    out = tmp_path / "ipmac.xlsx"
    write_ip_mac_report_excel(_report(), str(out))
    col_a = [c for c in _cells(openpyxl.load_workbook(out).active) if c]
    assert "IP/MAC Bindings Found On These Switches" in col_a
    assert "VLAN 5: sw-01" in col_a
    assert "VLAN 2003: not found on any selected switch" in col_a


def test_detail_table_headers_and_a_row(tmp_path):
    from excel_generator import write_ip_mac_report_excel, IP_MAC_HEADERS
    out = tmp_path / "ipmac.xlsx"
    write_ip_mac_report_excel(_report(), str(out))
    ws = openpyxl.load_workbook(out).active

    hdr_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == "Switch")
    headers = [ws.cell(row=hdr_row, column=c).value for c in range(1, len(IP_MAC_HEADERS) + 1)]
    assert headers == IP_MAC_HEADERS

    first = [ws.cell(row=hdr_row + 1, column=c).value for c in range(1, len(IP_MAC_HEADERS) + 1)]
    assert first == ["sw-01", "2", "Fi2/0/32", "5", "10.0.5.10",
                     "0011.2233.aaaa", "REACHABLE", "145s", None]


def test_flagged_rows_are_highlighted(tmp_path):
    from excel_generator import write_ip_mac_report_excel, IP_MAC_FLAG_COLOUR
    out = tmp_path / "ipmac.xlsx"
    write_ip_mac_report_excel(_report(), str(out))
    ws = openpyxl.load_workbook(out).active

    hdr_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == "Switch")
    unflagged = ws.cell(row=hdr_row + 1, column=1)
    flagged = ws.cell(row=hdr_row + 2, column=1)
    assert IP_MAC_FLAG_COLOUR in str(flagged.fill.start_color.rgb)
    assert flagged.font.bold is True
    assert IP_MAC_FLAG_COLOUR not in str(unflagged.fill.start_color.rgb)


def test_exclusion_lines_appear_only_when_non_zero(tmp_path):
    from excel_generator import write_ip_mac_report_excel
    clean = tmp_path / "clean.xlsx"
    write_ip_mac_report_excel(_report(), str(clean))
    col_a = " ".join(c for c in _cells(openpyxl.load_workbook(clean).active) if c)
    assert "local/static" not in col_a
    assert "non-IPv4" not in col_a
    assert "did not support" not in col_a

    noisy = tmp_path / "noisy.xlsx"
    write_ip_mac_report_excel(
        _report(excluded_local=4, non_ipv4=2,
                unsupported=["sw-legacy-01"], no_bindings=["sw-03"]),
        str(noisy))
    col_a = " ".join(c for c in _cells(openpyxl.load_workbook(noisy).active) if c)
    assert "Excluded 4 local/static row(s)" in col_a
    assert "Skipped 2 non-IPv4 binding(s)" in col_a
    assert "did not support the command: sw-legacy-01" in col_a
    assert "returned no bindings: sw-03" in col_a


def test_refuses_to_write_an_empty_report(tmp_path):
    from excel_generator import write_ip_mac_report_excel
    out = tmp_path / "empty.xlsx"
    ok, msg = write_ip_mac_report_excel(_report(rows=[]), str(out))
    assert ok is False
    assert not out.exists()
    assert "no" in msg.lower()
