import openpyxl

from excel_generator import write_excel_sheet
from interface_parser import InterfaceRecord


def test_link_change_column_absent_when_no_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    rec = InterfaceRecord(switch="sw-a", iface="Gi1/0/1", state="connected")
    write_excel_sheet(ws, [rec], {})
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Last Link Change" not in headers


def test_link_change_column_present_and_populated():
    wb = openpyxl.Workbook()
    ws = wb.active
    rec = InterfaceRecord(
        switch="sw-a", iface="Gi1/0/1", state="connected", last_link_change="2h13m"
    )
    write_excel_sheet(ws, [rec], {})
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers[-1] == "Last Link Change"
    last_col = ws.max_column
    assert ws.cell(row=2, column=last_col).value == "2h13m"


def test_end_to_end_link_change_flows_to_combined_workbook(tmp_path):
    from interface_parser import parse_output
    from excel_generator import write_combined_excel

    text = "\n".join([
        "GigabitEthernet1/0/5 is up, line protocol is up (connected)",
        "  Last input 00:00:01, output 00:00:00, output hang never",
        "*12:00:00.000 AEST Sun Jul 6 2026",
        "Log Buffer (16384 bytes):",
        "*Jul  6 09:47:00.000: %LINK-3-UPDOWN: Interface GigabitEthernet1/0/5, changed state to up",
    ])
    records, stack = parse_output(text, "hu-chi-f1-edge-01")
    assert len(records) == 1  # fail loudly here if parsing regresses
    devices = {"hu-chi-f1-edge-01": (records, stack)}

    out = tmp_path / "report.xlsx"
    ok, _ = write_combined_excel(devices, 42, str(out))
    assert ok

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["All Ports"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers[-1] == "Last Link Change"
    assert ws.cell(row=2, column=ws.max_column).value == "2h13m"
