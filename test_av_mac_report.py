# test_av_mac_report.py
HIERARCHY_OUTPUTS = {
    "acc1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    0011.2233.4455    DYNAMIC     Gi1/0/24",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Device ID: dist1",
        "Entry address(es):",
        "  IP address: 10.0.0.2",
        "Platform: cisco WS-C3850-48P,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/0/1,  Port ID (outgoing port): GigabitEthernet1/0/1",
        "Total cdp entries displayed : 1",
    ]),
    "dist1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    0011.2233.4455    DYNAMIC     Gi1/0/1",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Device ID: acc1",
        "Entry address(es):",
        "  IP address: 10.0.0.1",
        "Platform: cisco WS-C3560X-24,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/0/1,  Port ID (outgoing port): GigabitEthernet1/0/24",
        "-------------------------",
        "Device ID: core1",
        "Entry address(es):",
        "  IP address: 10.0.0.3",
        "Platform: cisco WS-C4500X-32,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/0/2,  Port ID (outgoing port): GigabitEthernet1/1/1",
        "Total cdp entries displayed : 2",
    ]),
    "core1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    0011.2233.4455    DYNAMIC     Gi1/1/1",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Device ID: dist1",
        "Entry address(es):",
        "  IP address: 10.0.0.2",
        "Platform: cisco WS-C3850-48P,  Capabilities: Router Switch IGMP",
        "Interface: GigabitEthernet1/1/1,  Port ID (outgoing port): GigabitEthernet1/0/2",
        "Total cdp entries displayed : 1",
    ]),
}


def test_build_av_mac_report_collapses_hierarchy_duplicates():
    from av_mac_report import build_av_mac_report
    report = build_av_mac_report(HIERARCHY_OUTPUTS, ["900"])
    assert len(report.rows) == 1
    row = report.rows[0]
    assert row.switch == "acc1"
    assert row.interface == "Gi1/0/24"
    assert row.mac == "0011.2233.4455"
    assert row.notes == ""


MULTI_MAC_OUTPUTS = {
    "acc1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    aaaa.aaaa.aaaa    DYNAMIC     Gi1/0/5",
        " 900    bbbb.bbbb.bbbb    DYNAMIC     Gi1/0/5",
        "Total Mac Addresses for this criterion: 2",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Total cdp entries displayed : 0",
    ]),
}


def test_build_av_mac_report_flags_multiple_macs_on_one_port():
    from av_mac_report import build_av_mac_report
    report = build_av_mac_report(MULTI_MAC_OUTPUTS, ["900"])
    assert len(report.rows) == 2
    assert all("Multiple MACs" in r.notes for r in report.rows)


AMBIGUOUS_OUTPUTS = {
    "acc1": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    cccc.cccc.cccc    DYNAMIC     Gi1/0/8",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Total cdp entries displayed : 0",
    ]),
    "acc2": "\n".join([
        "show mac address-table",
        "Vlan    Mac Address       Type        Ports",
        "----    -----------       --------    -----",
        " 900    cccc.cccc.cccc    DYNAMIC     Gi1/0/9",
        "Total Mac Addresses for this criterion: 1",
        "",
        "show cdp neighbors detail",
        "-------------------------",
        "Total cdp entries displayed : 0",
    ]),
}


def test_build_av_mac_report_flags_ambiguous_mac_across_switches():
    from av_mac_report import build_av_mac_report
    report = build_av_mac_report(AMBIGUOUS_OUTPUTS, ["900"])
    assert len(report.rows) == 2
    assert all("Ambiguous" in r.notes for r in report.rows)


def test_build_av_mac_report_filters_requested_vlans_and_builds_summary():
    from av_mac_report import build_av_mac_report
    outputs = {
        "acc1": "\n".join([
            "show mac address-table",
            "Vlan    Mac Address       Type        Ports",
            "----    -----------       --------    -----",
            " 900    1111.1111.1111    DYNAMIC     Gi1/0/1",
            " 905    2222.2222.2222    DYNAMIC     Gi1/0/2",
            " 10     3333.3333.3333    DYNAMIC     Gi1/0/3",
            "Total Mac Addresses for this criterion: 3",
            "",
            "show cdp neighbors detail",
            "-------------------------",
            "Total cdp entries displayed : 0",
        ]),
    }
    report = build_av_mac_report(outputs, ["900", "905"])
    assert sorted(r.vlan for r in report.rows) == ["900", "905"]
    assert report.vlan_stacks == {"900": ["acc1"], "905": ["acc1"]}


def test_build_av_mac_report_no_matching_vlan_returns_no_rows():
    from av_mac_report import build_av_mac_report
    report = build_av_mac_report(MULTI_MAC_OUTPUTS, ["999"])
    assert report.rows == []
    assert report.vlan_stacks == {"999": []}


def test_write_av_mac_report_excel_writes_summary_and_detail(tmp_path):
    import openpyxl
    from av_mac_report import AvMacReport, AvMacRow
    from excel_generator import write_av_mac_report_excel

    report = AvMacReport(
        vlan_stacks={"900": ["acc1"]},
        vlan_counts={"900": 2},
        rows=[
            AvMacRow(switch="acc1", stack_member="1", interface="Gi1/0/24",
                     vlan="900", mac="0011.2233.4455", type="DYNAMIC", notes=""),
            AvMacRow(switch="acc1", stack_member="1", interface="Gi1/0/5",
                     vlan="900", mac="aaaa.aaaa.aaaa", type="DYNAMIC",
                     notes="Multiple MACs — possible unmanaged switch"),
        ],
    )
    out = tmp_path / "av.xlsx"
    ok, msg = write_av_mac_report_excel(report, str(out))
    assert ok

    ws = openpyxl.load_workbook(out)["AV MAC-Port Export"]
    assert ws.cell(row=1, column=1).value == "AV VLANs Found On These Switches"
    assert ws.cell(row=2, column=1).value == "VLAN 900: acc1 (2 MAC(s))"

    header_row = next(r for r in range(1, ws.max_row + 1)
                       if ws.cell(row=r, column=1).value == "Switch")
    assert [ws.cell(row=header_row, column=c).value for c in range(1, 9)] == [
        "Switch", "Stack Member", "Interface", "VLAN", "MAC Address", "Type",
        "Device Type", "Notes"
    ]
    assert ws.cell(row=header_row + 1, column=5).value == "0011.2233.4455"
    assert ws.cell(row=header_row + 2, column=8).value == "Multiple MACs — possible unmanaged switch"


def test_write_av_mac_report_excel_fails_on_empty_rows(tmp_path):
    from av_mac_report import AvMacReport
    from excel_generator import write_av_mac_report_excel
    report = AvMacReport(vlan_stacks={}, rows=[])
    ok, msg = write_av_mac_report_excel(report, str(tmp_path / "av.xlsx"))
    assert ok is False
    assert "No matching" in msg
